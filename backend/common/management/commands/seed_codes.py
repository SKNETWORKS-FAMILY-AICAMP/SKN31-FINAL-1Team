import os
import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
from common.models import CommonCodeGroup, CommonCode


class Command(BaseCommand):
    help = "DB_명세서.xlsx의 초기 데이터(Seed) 시트 데이터를 DB에 자동 주입합니다."

    def handle(self, *args, **options):
        excel_path = os.path.join(settings.BASE_DIR, "DB_table.xlsx")

        if not os.path.exists(excel_path):
            self.stderr.write(self.style.ERROR(f"파일을 찾을 수 없습니다: {excel_path}"))
            return

        wb = openpyxl.load_workbook(excel_path)
        if "초기 데이터(Seed)" not in wb.sheetnames:
            self.stderr.write(self.style.ERROR("'초기 데이터(Seed)' 시트가 엑셀에 존재하지 않습니다."))
            return

        sheet = wb["초기 데이터(Seed)"]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            self.stderr.write(self.style.WARNING("엑셀 데이터가 비어 있습니다."))
            return

        # 헤더 제외한 데이터 행
        header = rows[0]
        data_rows = rows[1:]

        group_count = 0
        code_count = 0

        self.stdout.write("시드 데이터 주입을 시작합니다...")

        for row in data_rows:
            if not row or len(row) < 4:
                continue

            big_category, mid_category, sub_category, code_name = row[:4]

            if not mid_category or not sub_category or not code_name:
                continue

            # 코드 그룹 ID 생성 (예: USER_DEPARTMENT, SKILL_BACKEND 등)
            group_code = f"{big_category}_{mid_category}".upper()
            group_name = f"{big_category} - {mid_category}"

            # 1. CommonCodeGroup 생성 또는 조회
            group_obj, created_group = CommonCodeGroup.objects.get_or_create(
                group_code=group_code,
                defaults={
                    "group_name": group_name,
                    "description": f"{big_category} 대분류 하위 {mid_category} 그룹",
                    "is_active": True,
                }
            )
            if created_group:
                group_count += 1

            # 2. CommonCode 생성 또는 업데이트
            code_id = str(sub_category).strip()
            code_obj, created_code = CommonCode.objects.update_or_create(
                code_id=code_id,
                group=group_obj,
                defaults={
                    "code_name": str(code_name).strip(),
                    "is_active": True,
                }
            )
            if created_code:
                code_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"시드 데이터 주입 완료! (새로 생성된 코드 그룹: {group_count}개, 상세 코드: {code_count}개)"
            )
        )